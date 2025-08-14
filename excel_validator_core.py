import os
import tempfile
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Any, Optional
import re
from datetime import datetime, timedelta
from validate_email import validate_email
import pycountry

class ValidationError:
    """Classe pour représenter une erreur de validation"""
    def __init__(self, row: int, columns: List[str], message: str, values: List[Any] = None):
        self.row = row
        self.columns = columns if isinstance(columns, list) else [columns]
        self.message = message
        self.values = values or []
        self.coordinate = "+".join([f"{col}{row}" for col in self.columns])
        # Pour compatibilité avec l'ancienne structure
        self.column = self.columns[0] if self.columns else "A"
        self.value = self.values[0] if self.values else None

class ExcelValidatorCore:
    """Validateur Excel principal avec support multicolonne"""
    
    def __init__(self):
        self.errors: List[ValidationError] = []
        self.worksheet_data = {}  # Cache des données pour les règles conditionnelles
        self.multicolumn_cache = {}  # Cache pour les combinaisons de colonnes
    
    def validate_file(self, file_path: str, rules_config: Dict[str, Any], 
                     sheet_name: str = None) -> Tuple[bool, List[ValidationError], str]:
        """Valide un fichier Excel selon les règles fournies"""
        self.errors = []
        self.worksheet_data = {}
        self.multicolumn_cache = {}
        
        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Charger toutes les données en mémoire
            self._load_worksheet_data(ws)
            
            validators = rules_config.get("validators", {}).get("columns", {})
            default_validator = rules_config.get("validators", {}).get("default")
            excludes = rules_config.get("excludes", [])
            header_row = rules_config.get("header", True)
            conditional_rules = rules_config.get("conditional_rules", [])
            multicolumn_rules = rules_config.get("multicolumn_rules", [])  
            multi_simple_rules = rules_config.get("multi_simple_rules", [])

            # Validation des règles simples
            self._validate_worksheet(ws, validators, default_validator, excludes, header_row)
            
            # Validation des règles simples multicolonnes
            self._validate_multi_simple_rules(multi_simple_rules)
            
            # Validation des règles conditionnelles
            self._validate_conditional_rules(conditional_rules)
            
            # Validation des règles multicolonnes
            self._validate_multicolumn_rules(multicolumn_rules)
            
            error_file_path = None
            if self.errors:
                error_file_path = self._generate_error_file(file_path, ws, self.errors)
            
            wb.close()
            return len(self.errors) == 0, self.errors, error_file_path
            
        except Exception as e:
            error = ValidationError(0, ["A"], f"Erreur lors de la lecture du fichier: {str(e)}")
            return False, [error], None
    
    def _validate_multi_simple_rules(self, multi_simple_rules: List[Dict]):
        """Valide les règles simples appliquées à plusieurs colonnes"""
        for rule in multi_simple_rules:
            rule_id = rule.get("id", "unknown")
            columns = rule.get("columns", [])
            rule_type = rule.get("rule_type", "")
            params = rule.get("params", {})
            message = rule.get("message", f"Erreur règle simple multicolonne {rule_type}")
            
            # Valider chaque ligne pour chaque colonne concernée
            for row_idx, row_data in self.worksheet_data.items():
                if row_idx == 1:  # Skip header
                    continue
                
                # Vérifier chaque colonne de la règle
                for column in columns:
                    value = row_data.get(column)
                    
                    # Appliquer la même logique que pour les règles simples
                    is_valid = self._validate_simple_rule_value(value, rule_type, params)
                    
                    if not is_valid:
                        error = ValidationError(row_idx, [column], message, [value])
                        self.errors.append(error) 

    def _validate_simple_rule_value(self, value: Any, rule_type: str, params: Dict) -> bool:
        """Valide une valeur selon une règle simple (réutilise la logique existante)"""
        try:
            # Appliquer le trim si nécessaire
            if params.get("trim", False) and isinstance(value, str):
                value = value.strip()
            
            if rule_type == "NotBlank":
                return self._validate_not_blank(value, params)
            elif rule_type == "Length":
                return self._validate_length(value, params)
            elif rule_type == "Type":
                return self._validate_type(value, params)
            elif rule_type == "Regex":
                return self._validate_regex(value, params)
            elif rule_type == "Email":
                return self._validate_email(value, params)
            elif rule_type == "Choice":
                return self._validate_choice(value, params)
            elif rule_type == "Country":
                return self._validate_country(value, params)
            elif rule_type == "Date" or rule_type == "ExcelDate":
                return self._validate_date(value, params)
            elif rule_type == "Comparison":
                return self._validate_comparison(value, params)
            elif rule_type == "Duplicate":
                # Pour les doublons sur plusieurs colonnes, on ne peut pas utiliser la même logique
                # On pourrait implémenter une logique spécifique si nécessaire
                return True
            
            return True
            
        except Exception as e:
            print(f"Erreur validation règle simple {rule_type}: {e}")
            return False

    def _validate_multicolumn_rules(self, multicolumn_rules: List[Dict]):
        """Valide les règles multicolonnes"""
        for rule in multicolumn_rules:
            rule_id = rule.get("id", "unknown")
            columns = rule.get("columns", [])
            rule_type = rule.get("rule_type", "")
            params = rule.get("params", {})
            message = rule.get("message", f"Erreur règle multicolonne {rule_type}")
            
            # Valider chaque ligne
            for row_idx, row_data in self.worksheet_data.items():
                if row_idx == 1:  # Skip header
                    continue
                    
                # Récupérer les valeurs des colonnes concernées
                values = []
                for col in columns:
                    values.append(row_data.get(col))
                
                # Appliquer la validation selon le type de règle
                is_valid = self._validate_multicolumn_rule(rule_type, columns, values, params, row_idx)
                
                if not is_valid:
                    error = ValidationError(row_idx, columns, message, values)
                    self.errors.append(error)
    
    def _validate_multicolumn_rule(self, rule_type: str, columns: List[str], values: List[Any], 
                                 params: Dict, row_idx: int) -> bool:
        """Valide une règle multicolonne spécifique"""
        try:
            if rule_type == "sum_equals":
                return self._validate_sum_equals(values, params)
            elif rule_type == "sum_range":
                return self._validate_sum_range(values, params)
            elif rule_type == "date_before":
                return self._validate_date_before(values)
            elif rule_type == "date_after":
                return self._validate_date_after(values)
            elif rule_type == "date_range":
                return self._validate_date_range(values, params)
            elif rule_type == "percentage_of":
                return self._validate_percentage_of(values, params)
            elif rule_type == "all_or_none":
                return self._validate_all_or_none(values)
            elif rule_type == "unique_combination":
                return self._validate_unique_combination(columns, values, params, row_idx)
            elif rule_type == "conditional_sum":
                return self._validate_conditional_sum(columns, values, params, row_idx)
            elif rule_type == "max_min_check":
                return self._validate_max_min_check(values, params)
            
            return True
            
        except Exception as e:
            print(f"Erreur validation multicolonne {rule_type}: {e}")
            return False
    
    def _validate_sum_equals(self, values: List[Any], params: Dict) -> bool:
        """Valide que la somme des colonnes source égale la colonne cible"""
        try:
            target_column = params.get("target_column", "")
            if not target_column:
                return True
                
            # Les colonnes source sont toutes sauf la cible
            source_values = values[:-1]  # Toutes sauf la dernière (cible)
            target_value = values[-1]    # La dernière est la cible
            
            # Convertir en nombres
            source_nums = []
            for val in source_values:
                if val is None or val == "":
                    source_nums.append(0)
                else:
                    source_nums.append(float(val))
            
            if target_value is None or target_value == "":
                target_num = 0
            else:
                target_num = float(target_value)
            
            calculated_sum = sum(source_nums)
            tolerance = params.get("tolerance", 0.01)
            
            return abs(calculated_sum - target_num) <= tolerance
            
        except (ValueError, TypeError):
            return False
    
    def _validate_sum_range(self, values: List[Any], params: Dict) -> bool:
        """Valide que la somme des colonnes est dans une plage"""
        try:
            min_value = params.get("min_value", 0)
            max_value = params.get("max_value", float('inf'))
            
            nums = []
            for val in values:
                if val is None or val == "":
                    nums.append(0)
                else:
                    nums.append(float(val))
            
            total_sum = sum(nums)
            return min_value <= total_sum <= max_value
            
        except (ValueError, TypeError):
            return False
    
    def _validate_date_before(self, values: List[Any]) -> bool:
        """Valide que la première date est antérieure à la seconde"""
        if len(values) < 2:
            return True
            
        try:
            date1 = self._parse_date(values[0])
            date2 = self._parse_date(values[1])
            
            if date1 is None or date2 is None:
                return True  # Ne pas valider si les dates sont manquantes
                
            return date1 < date2
            
        except Exception:
            return False
    
    def _validate_date_after(self, values: List[Any]) -> bool:
        """Valide que la première date est postérieure à la seconde"""
        if len(values) < 2:
            return True
            
        try:
            date1 = self._parse_date(values[0])
            date2 = self._parse_date(values[1])
            
            if date1 is None or date2 is None:
                return True
                
            return date1 > date2
            
        except Exception:
            return False
    
    def _validate_date_range(self, values: List[Any], params: Dict) -> bool:
        """Valide que l'écart entre deux dates est dans une plage"""
        if len(values) < 2:
            return True
            
        try:
            date1 = self._parse_date(values[0])
            date2 = self._parse_date(values[1])
            
            if date1 is None or date2 is None:
                return True
                
            min_days = params.get("min_days", 0)
            max_days = params.get("max_days", 365)
            
            diff = abs((date2 - date1).days)
            return min_days <= diff <= max_days
            
        except Exception:
            return False
    
    def _validate_percentage_of(self, values: List[Any], params: Dict) -> bool:
        """Valide qu'une valeur représente un pourcentage d'une autre"""
        if len(values) < 2:
            return True
            
        try:
            value1 = float(values[0]) if values[0] not in [None, ""] else 0
            value2 = float(values[1]) if values[1] not in [None, ""] else 0
            
            if value2 == 0:
                return value1 == 0
                
            expected_percentage = params.get("percentage", 0) / 100.0
            tolerance = params.get("tolerance", 0.05)  # 5% de tolérance par défaut
            
            actual_percentage = value1 / value2
            expected_value = value2 * expected_percentage
            
            return abs(value1 - expected_value) <= (value2 * tolerance)
            
        except (ValueError, TypeError, ZeroDivisionError):
            return False
    
    def _validate_all_or_none(self, values: List[Any]) -> bool:
        """Valide que toutes les valeurs sont remplies ou toutes vides"""
        non_empty_count = 0
        for val in values:
            if val is not None and str(val).strip() != "":
                non_empty_count += 1
        
        # Soit toutes vides (0) soit toutes remplies (len(values))
        return non_empty_count == 0 or non_empty_count == len(values)
    
    def _validate_unique_combination(self, columns: List[str], values: List[Any], 
                                   params: Dict, current_row: int) -> bool:
        """Valide que la combinaison des valeurs est unique"""
        case_sensitive = params.get("case_sensitive", True)
        
        # Créer la clé de combinaison
        combination_key = tuple(columns)
        if combination_key not in self.multicolumn_cache:
            self.multicolumn_cache[combination_key] = {}
        
        # Normaliser les valeurs
        normalized_values = []
        for val in values:
            if val is None:
                normalized_values.append("")
            else:
                str_val = str(val)
                if not case_sensitive:
                    str_val = str_val.lower()
                normalized_values.append(str_val)
        
        combination_value = tuple(normalized_values)
        
        # Vérifier l'unicité
        for row_idx, row_data in self.worksheet_data.items():
            if row_idx == current_row or row_idx == 1:  # Skip current row and header
                continue
            
            other_values = []
            for col in columns:
                other_val = row_data.get(col)
                if other_val is None:
                    other_values.append("")
                else:
                    str_val = str(other_val)
                    if not case_sensitive:
                        str_val = str_val.lower()
                    other_values.append(str_val)
            
            other_combination = tuple(other_values)
            
            if combination_value == other_combination:
                return False
        
        return True
    
    def _validate_conditional_sum(self, columns: List[str], values: List[Any], 
                                params: Dict, row_idx: int) -> bool:
        """Valide une somme conditionnelle"""
        condition_column = params.get("condition_column", "")
        condition_value = params.get("condition_value", "")
        operator = params.get("operator", "greater_than")
        target_value = params.get("target_value", 0)
        
        if not condition_column:
            return True
            
        # Vérifier la condition
        condition_cell_value = self.worksheet_data.get(row_idx, {}).get(condition_column)
        
        if str(condition_cell_value) != str(condition_value):
            return True  # Condition non remplie, pas d'erreur
        
        # Calculer la somme
        try:
            nums = []
            for val in values:
                if val is None or val == "":
                    nums.append(0)
                else:
                    nums.append(float(val))
            
            total_sum = sum(nums)
            target_num = float(target_value)
            
            if operator == "greater_than":
                return total_sum > target_num
            elif operator == "less_than":
                return total_sum < target_num
            elif operator == "equals":
                return abs(total_sum - target_num) <= 0.01
            elif operator == "greater_equal":
                return total_sum >= target_num
            elif operator == "less_equal":
                return total_sum <= target_num
            
        except (ValueError, TypeError):
            return False
            
        return True
    
    def _validate_max_min_check(self, values: List[Any], params: Dict) -> bool:
        """Valide qu'une colonne contient le max/min des autres"""
        operation = params.get("operation", "max").lower()
        target_column = params.get("target_column", "last")
        
        if len(values) < 2:
            return True
            
        try:
            # Séparer la valeur cible des valeurs source
            if target_column == "last":
                source_values = values[:-1]
                target_value = values[-1]
            elif target_column == "first":
                source_values = values[1:]
                target_value = values[0]
            else:
                # Utiliser toutes les valeurs comme source et cible
                source_values = values
                target_value = values[0]
            
            # Convertir en nombres
            source_nums = []
            for val in source_values:
                if val is not None and val != "":
                    source_nums.append(float(val))
            
            if not source_nums:
                return True
                
            if target_value is None or target_value == "":
                return False
                
            target_num = float(target_value)
            
            if operation == "max":
                expected = max(source_nums)
            elif operation == "min":
                expected = min(source_nums)
            else:
                return True
                
            tolerance = params.get("tolerance", 0.01)
            return abs(target_num - expected) <= tolerance
            
        except (ValueError, TypeError):
            return False
    
    def _parse_date(self, value: Any) -> Optional[datetime]:
        """Parse une valeur en date"""
        if value is None or value == "":
            return None
            
        if isinstance(value, datetime):
            return value
            
        if isinstance(value, str):
            # Essayer différents formats de date
            date_formats = [
                "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", 
                "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"
            ]
            
            for fmt in date_formats:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
                    
        return None
    
    def _load_worksheet_data(self, ws):
        """Charge toutes les données de la feuille en mémoire"""
        self.header_map = {}  # Dictionnaire Colonne -> Nom
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if all(cell is None or cell == "" for cell in row):
                continue
            self.worksheet_data[row_idx] = {}
            for col_idx, value in enumerate(row, 1):
                column_letter = get_column_letter(col_idx)
                self.worksheet_data[row_idx][column_letter] = value
                # Stocker le nom de l'en-tête si on est sur la première ligne
                if row_idx == 1 and value not in (None, ""):
                    self.header_map[column_letter] = str(value)
    
    def _validate_conditional_rules(self, conditional_rules: List[Dict]):
        """Valide les règles conditionnelles"""
        for rule in conditional_rules:
            if not rule.get("active", True):
                continue
                
            conditions = rule.get("conditions", [])
            actions = rule.get("actions", [])
            logic = rule.get("logic", "AND")
            message = rule.get("message", "Règle conditionnelle non respectée")
            
            for row_idx, row_data in self.worksheet_data.items():
                if row_idx == 1:  # Skip header
                    continue
                    
                # Évaluer les conditions
                condition_results = []
                for condition in conditions:
                    column = condition["column"]
                    operator = condition["operator"]
                    value = condition.get("value", "")
                    
                    cell_value = row_data.get(column)
                    condition_result = self._evaluate_condition(cell_value, operator, value)
                    condition_results.append(condition_result)
                
                # Combiner les résultats selon la logique
                if logic == "AND":
                    conditions_met = all(condition_results)
                else:  # OR
                    conditions_met = any(condition_results)
                
                # Si conditions remplies, vérifier les actions
                if conditions_met:
                    for action in actions:
                        action_column = action["column"]
                        action_type = action["type"]
                        action_params = action.get("params", {})
                        
                        cell_value = row_data.get(action_column)
                        if not self._validate_action(cell_value, action_type, action_params):
                            error = ValidationError(row_idx, [action_column], message, [cell_value])
                            self.errors.append(error)
    
    def _evaluate_condition(self, cell_value: Any, operator: str, compare_value: str) -> bool:
        """Évalue une condition"""
        if cell_value is None:
            cell_value = ""
        
        cell_str = str(cell_value).strip()
        compare_str = str(compare_value).strip()
        
        try:
            if operator == "equals":
                return cell_str == compare_str
            elif operator == "not_equals":
                return cell_str != compare_str
            elif operator == "greater_than":
                return float(cell_value) > float(compare_value)
            elif operator == "less_than":
                return float(cell_value) < float(compare_value)
            elif operator == "greater_equal":
                return float(cell_value) >= float(compare_value)
            elif operator == "less_equal":
                return float(cell_value) <= float(compare_value)
            elif operator == "starts_with":
                return cell_str.startswith(compare_str)
            elif operator == "ends_with":
                return cell_str.endswith(compare_str)
            elif operator == "contains":
                return compare_str in cell_str
            elif operator == "not_contains":
                return compare_str not in cell_str
            elif operator == "is_empty":
                return cell_str == ""
            elif operator == "is_not_empty":
                return cell_str != ""
        except (ValueError, TypeError):
            return False
        
        return False
    
    def _validate_action(self, value: Any, action_type: str, params: Dict) -> bool:
        """Valide une action conditionnelle"""
        if action_type == "must_be_empty":
            return value is None or str(value).strip() == ""
        elif action_type == "must_not_be_empty":
            return value is not None and str(value).strip() != ""
        elif action_type == "must_be_between":
            try:
                val = float(value)
                min_val = params.get("min", 0)
                max_val = params.get("max", 100)
                return min_val <= val <= max_val
            except (ValueError, TypeError):
                return False
        elif action_type == "must_be_in_list":
            values_list = params.get("values", [])
            return str(value) in [str(v) for v in values_list]
        elif action_type == "must_match_pattern":
            pattern = params.get("pattern", "")
            try:
                return bool(re.match(pattern, str(value)))
            except re.error:
                return False
        
        return True
    
    def _validate_worksheet(self, ws, validators: Dict, default_validator: Optional[Dict], 
                          excludes: List[str], header_row: Any):
        """Valide une feuille de calcul avec les règles simples"""
        header_found = header_row is True
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if all(cell is None or cell == "" for cell in row):
                continue
            
            if not header_found and header_row != True:
                if header_row in row:
                    header_found = True
                continue
            
            for col_idx, value in enumerate(row, 1):
                column_letter = get_column_letter(col_idx)
                
                if column_letter in excludes:
                    continue
                
                if column_letter in validators:
                    for rule in validators[column_letter]:
                        self._apply_validation_rule(rule, value, row_idx, column_letter)
                elif default_validator:
                    for rule in default_validator:
                        self._apply_validation_rule(rule, value, row_idx, column_letter)
    
    def _apply_validation_rule(self, rule: Dict, value: Any, row: int, column: str):
        """Applique une règle de validation à une cellule"""
        rule_type = list(rule.keys())[0]
        rule_params = rule[rule_type]
        
        try:
            if rule_params.get("trim", False) and isinstance(value, str):
                value = value.strip()
            
            is_valid = True
            error_message = rule_params.get("message", f"Erreur de validation {rule_type}")
            
            if rule_type == "NotBlank":
                is_valid = self._validate_not_blank(value, rule_params)
            elif rule_type == "Length":
                is_valid = self._validate_length(value, rule_params)
            elif rule_type == "Type":
                is_valid = self._validate_type(value, rule_params)
            elif rule_type == "Regex":
                is_valid = self._validate_regex(value, rule_params)
            elif rule_type == "Email":
                is_valid = self._validate_email(value, rule_params)
            elif rule_type == "Choice":
                is_valid = self._validate_choice(value, rule_params)
            elif rule_type == "Country":
                is_valid = self._validate_country(value, rule_params)
            elif rule_type == "Date" or rule_type == "ExcelDate":
                is_valid = self._validate_date(value, rule_params)
            elif rule_type == "Comparison":
                is_valid = self._validate_comparison(value, rule_params)
            elif rule_type == "Duplicate":
                is_valid = self._validate_duplicate(value, column, row, rule_params)
            
            if not is_valid:
                error = ValidationError(row, [column], error_message, [value])
                self.errors.append(error)
                
        except Exception as e:
            error = ValidationError(row, [column], f"Erreur lors de la validation: {str(e)}", [value])
            self.errors.append(error)
    
    def _validate_not_blank(self, value: Any, params: Dict) -> bool:
        """Validation NotBlank"""
        return value is not None and str(value).strip() != ""
    
    def _validate_length(self, value: Any, params: Dict) -> bool:
        """Validation Length"""
        if value is None:
            return True
        str_value = str(value)
        min_length = params.get("min")
        max_length = params.get("max")
        
        if min_length is not None and len(str_value) < min_length:
            return False
        if max_length is not None and len(str_value) > max_length:
            return False
        return True
    
    def _validate_type(self, value: Any, params: Dict) -> bool:
        """Validation Type"""
        if value is None:
            return True
        expected_type = params.get("type", "").lower()
        
        try:
            if expected_type == "integer":
                int(value)
                return True
            elif expected_type == "float":
                float(value)
                return True
            elif expected_type == "bool":
                return str(value) in ["0", "1", "True", "False", "true", "false"]
        except (ValueError, TypeError):
            return False
        return True
    
    def _validate_regex(self, value: Any, params: Dict) -> bool:
        """Validation Regex"""
        if value is None:
            return True
        pattern = params.get("pattern")
        if not pattern:
            return True
        try:
            return bool(re.match(pattern, str(value)))
        except re.error:
            return False
    
    def _validate_email(self, value: Any, params: Dict) -> bool:
        """Validation Email"""
        if value is None:
            return True
        if not isinstance(value, str):
            return False
        try:
            return validate_email(value)
        except:
            return False
    
    def _validate_choice(self, value: Any, params: Dict) -> bool:
        """Validation Choice"""
        if value is None:
            return True
        choices = params.get("choices", [])
        case_sensitive = params.get("caseSensitive", True)
        
        str_value = str(value)
        if not case_sensitive:
            str_value = str_value.lower()
            choices = [str(choice).lower() for choice in choices]
        
        return str_value in choices
    
    def _validate_country(self, value: Any, params: Dict) -> bool:
        """Validation Country"""
        if value is None:
            return True
        try:
            pycountry.countries.get(name=str(value))
            return True
        except (KeyError, LookupError):
            return False
    
    def _validate_date(self, value: Any, params: Dict) -> bool:
        """Validation Date"""
        if value is None:
            return True
        if isinstance(value, datetime):
            return True
        if isinstance(value, str):
            date_format = params.get("format", "%Y-%m-%d")
            try:
                datetime.strptime(value, date_format)
                return True
            except ValueError:
                return False
        return False
    
    def _validate_comparison(self, value: Any, params: Dict) -> bool:
        """Validation Comparison"""
        if value is None:
            return True
        
        operator = params.get("operator", "equals")
        compare_value = params.get("value", "")
        
        return self._evaluate_condition(value, operator, compare_value)
    
    def _validate_duplicate(self, value: Any, column: str, current_row: int, params: Dict) -> bool:
        """Validation Duplicate"""
        if value is None or str(value).strip() == "":
            return True
        
        case_sensitive = params.get("caseSensitive", True)
        check_value = str(value)
        if not case_sensitive:
            check_value = check_value.lower()
        
        # Compter les occurrences dans la colonne
        count = 0
        for row_idx, row_data in self.worksheet_data.items():
            if row_idx == current_row:
                continue
            
            cell_value = row_data.get(column)
            if cell_value is not None:
                compare_value = str(cell_value)
                if not case_sensitive:
                    compare_value = compare_value.lower()
                
                if compare_value == check_value:
                    count += 1
        
        return count == 0
    
    def _generate_error_file(self, original_file: str, ws, errors: List[ValidationError]) -> str:
        """Génère un fichier Excel avec les erreurs marquées"""
        try:
            timestamp = int(time.time())
            base_name = os.path.splitext(os.path.basename(original_file))[0]
            error_file = os.path.join(tempfile.gettempdir(), 
                                    f"errors_{timestamp}_{base_name}.xlsx")
            
            wb = load_workbook(original_file, data_only=True)
            ws = wb.active if ws is None else wb[ws.title]
            
            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
            
            for error in errors:
                try:
                    # Marquer toutes les cellules concernées
                    for column in error.columns:
                        cell = ws[f"{column}{error.row}"]
                        # Rouge pour erreurs simples, orange pour erreurs multicolonnes
                        if len(error.columns) > 1:
                            cell.fill = orange_fill
                        else:
                            cell.fill = red_fill
                except:
                    continue
            
            wb.save(error_file)
            wb.close()
            return error_file
        except Exception as e:
            print(f"Erreur génération fichier d'erreurs: {e}")
            return None
    
    def get_errors_as_dataframe(self) -> pd.DataFrame:
        """Retourne les erreurs sous forme de DataFrame"""
        if not self.errors:
            return pd.DataFrame(columns=["Ligne", "Colonne(s)", "Coordonnée", "Message", "Valeur(s)"])
        
        data = []
        # Récupérer la map des en-têtes une fois (vide si non chargée)
        header_map = getattr(self, "header_map", {}) or {}

        # Référence de la colonne B (cellule de référence demandée)
        ref_column = "B"

        def _col_label(col_letter: str, row: int) -> str:
            """Retourne 'NomColonne-B{row}(valeur)' si le nom existe, sinon 'ColLetter-B{row}(valeur)'.
            Récupère la valeur de la cellule de référence (colonne B) pour la ligne donnée.
            """
            name = header_map.get(col_letter)
            base = name if name else col_letter

            # Récupérer la valeur en colonne de référence pour cette ligne (si disponible)
            row_data = getattr(self, "worksheet_data", {}) or {}
            row_cells = row_data.get(row, {}) if isinstance(row_data, dict) else {}
            ref_val = row_cells.get(ref_column)
            ref_val_str = "" if ref_val is None else str(ref_val)

            # Format: 'value - Column' (ne pas afficher l'index de cellule). Si la valeur B est vide,
            # retourner seulement le nom/lettre de colonne.
            if ref_val_str:
                return f"{ref_val_str} - {base}"
            return base

        for error in self.errors:
            # Afficher 'NomColonne-B{row}' pour chaque colonne en erreur
            columns_str = ", ".join([_col_label(c, error.row) for c in error.columns])
            values_str = ", ".join([str(v) if v is not None else "" for v in error.values])
            
            data.append({
                "Ligne": error.row,
                "Colonne(s)": columns_str,
                "Coordonnée": error.coordinate,
                "Message": error.message,
                "Valeur(s)": values_str
            })
        
        return pd.DataFrame(data)
    
    def get_validation_summary(self) -> Dict[str, Any]:
        """Retourne un résumé de la validation"""
        if not self.errors:
            return {
                "status": "success",
                "total_errors": 0,
                "message": "✅ Aucune erreur détectée ! Le fichier est conforme aux règles."
            }
        
        error_by_type = {}
        multicolumn_errors = 0
        simple_errors = 0
        multi_simple_errors = 0
        
        for error in self.errors:
            if len(error.columns) > 1:
                multicolumn_errors += 1
                error_type = f"Erreur multicolonne ({len(error.columns)} colonnes)"
            elif "règle simple multicolonne" in error.message.lower():
                multi_simple_errors += 1
                error_type = "Erreur règle simple multicolonne"
            else:
                simple_errors += 1
                error_type = error.message.split(":")[0] if ":" in error.message else "Erreur simple"
            
            error_by_type[error_type] = error_by_type.get(error_type, 0) + 1
        
        return {
            "status": "error",
            "total_errors": len(self.errors),
            "simple_errors": simple_errors,
            "multi_simple_errors": multi_simple_errors,
            "multicolumn_errors": multicolumn_errors,
            "errors_by_type": error_by_type,
            "message": f"❌ {len(self.errors)} erreur(s) détectée(s) ({simple_errors} simples, {multicolumn_errors} multicolonnes)"
        }
